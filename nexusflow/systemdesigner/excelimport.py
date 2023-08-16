import logging
from typing import List, Any, Tuple
from openpyxl import load_workbook

from nexusflow.systemdesigner.module.pindefinition import PinDefinition, ConversionCoefficients, Gui, Value, \
    Miscellaneous, Exponential, Alarm

logging.basicConfig(level=logging.DEBUG)


ignore_rows = ('PWR',)


def get_display_type(function) -> Tuple[str, str]:
    if function == 'DIG_IN':
        return 'bool', 'output'
    elif function == 'DIG_OUT':
        return 'bool', 'input'
    elif function == 'ADC':
        return 'float', 'output'
    elif function == 'DAC':
        return 'float', 'input'
    else:
        raise ValueError(f'Unknown pin function {function}')


def try_import(row, column: int) -> Any:
    try:
        return row[column].value
    except IndexError:
        logging.info(f'Column {column} not found')
        return None


def import_excel(exel_file_path) -> List[PinDefinition]:
    excel_data = load_workbook(
        filename=exel_file_path,
        data_only=True,
        read_only=True
    ).worksheets[0]

    output_data = []

    for row in excel_data.iter_rows(min_row=6, max_row=excel_data.max_row, min_col=2, max_col=excel_data.max_column):
        if row[1].value not in ignore_rows and row[1].value is not None:
            pin_definition = PinDefinition(
                id=str(row[0].value),
                function=str(row[1].value),
                name=str(row[2].value),
                unit=str(row[3].value),
                conversion_coefficients=ConversionCoefficients(
                    k=float(row[4].value),
                    C=float(row[5].value)
                ),
                main_gui=Gui(
                    display=row[6].value == 0,
                    column=row[7].value,
                    row=row[8].value,
                    display_type=get_display_type(row[1].value)[0],
                    display_direction=get_display_type(row[1].value)[1]
                ),
                value=Value(
                    min=float(row[11].value),
                    max=float(row[12].value),
                    initial=float(row[13].value),
                    enable_initial=row[22].value == 1
                ),
                important_gui=Gui(
                    display=row[14].value == 1,
                    column=row[15].value,
                    row=row[16].value,
                    display_type=get_display_type(row[1].value)[0],
                    display_direction=get_display_type(row[1].value)[1]
                ),
                active_low=row[17].value == 1,
                exponential=Exponential(
                    enable=row[18].value == 1,
                    B=row[19].value,
                    R0=row[20].value,
                    T0=row[21].value
                ),
                alarm=Alarm(
                    enable=row[23].value == 1,
                    above=1 if row[24].value == 1 else 0,
                    warning_value=row[25].value,
                    interlock_value=row[26].value
                ),
                miscellaneous=Miscellaneous(
                    dac_range=row[9].value,
                    adc_range=row[10].value,
                    disable_powerdrop=try_import(row, 39)
                )
            )
            # logging.debug(pin_definition)
            output_data.append(pin_definition)

    # output_data.sort(key=lambda x: x.id)

    return output_data


if __name__ == '__main__':
    path = 'C:/Users/aspus/Desktop/ADD_adapter_test.xlsm'
    data = import_excel(path)
    for i in data:
        print(i)
